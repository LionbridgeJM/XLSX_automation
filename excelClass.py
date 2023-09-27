from email.mime import base
import openpyxl as xl
import xlsxwriter
from copy import copy
from openpyxl.utils.exceptions import *
import os
import re
from tkinter import messagebox
import os

import time


class ExcelProcesses:

    def __init__(self, path, opcion):
        self.path = path
        self.opcion = opcion
        self.mensajeretornado = ""

    def ftbt(self):
        numero = 0

        newpath = r'PrepFiles'
        if not os.path.exists(newpath):
            os.makedirs(newpath)
        wb1 = xl.load_workbook(self.path)
        ws1 = wb1.active

        rangeselected = []
        basename=os.path.basename(self.path)
     
        mr = ws1.max_row
        mc = ws1.max_column

        # FTBT DEFAULT PREP
        if self.opcion == 0:
            
            # Create an new Excel file xlsxwriter and add a worksheet.
            #workbook = xlsxwriter.Workbook('PrepFiles\\ '+ws1.title+".xlsx")
            workbook = xlsxwriter.Workbook('PrepFiles\\ '+basename)
            for ws1 in wb1.worksheets:
                numero += 1
                tab = str(numero)
                
                hoja = workbook.add_worksheet(ws1.title)

               
                #si yo le aumento a 120 si llena todo
                for i in range(1, ws1.max_row + 1):
                    if ws1.cell(row=i, column=6).value != None:
                        ws1.cell(row=i, column=7).value = ws1.cell(row=i, column=6).value
                        #Agregar esto en los otros metodos

                        textoModificado = re.sub(r"(?<=\w)(<br\/>)(?=\w)|(?<=\w) (<br\/>)(?=\w)|(?<=\w) (<br\/>) (?=\w)|(?<=\w)(<br\/>) (?=\w)|(<br\/>)(?=\()|(?<=\))(<br\/>)",
                                             "<LIONBRIDGE-br/>", str(ws1.cell(row=i, column=7).value))
                                             
                        ws1.cell(row=i, column=7).value = textoModificado

                        for j in range(1, mc + 2):
                            #print("Row: " + str (i-1) + " colum: " + str (j-1))
                            hoja.write(i-1, j-1, ws1.cell(row=i, column=j).value)
                hoja.set_column('A:F', None, None, {'hidden': 1})
                hoja.set_row(0, None, None, {'hidden': True})
                hoja.write(0, 5, 'Text')
                hoja.write(0, 6, 'Translation')
                hoja.write(0, 7, 'BackTranslation')
                hoja.write(0, 8, 'ComparativeReview')
            try:
                workbook.close()
                self.mensajeretornado = ("Successfully Done!")
                # ftbt.mensaje="Sucessfully completed"
                # tkinter.messagebox.showinfo("info","Sucessfully completed" )

                # return "Sucessfully completed"
            except xlsxwriter.exceptions.FileCreateError as e:
                numeto = 1
                self.mensajeretornado = "Error Close Current Prepared File"
                # return mgs
        elif self.opcion == 1:

            # Create an new Excel file xlsxwriter and add a worksheet.

            for ws1 in wb1.worksheets:
                numero += 1
                tab = str(numero)
                workbook = xlsxwriter.Workbook(
                    'PrepFiles\\ '+ws1.title+".xlsx")
                hoja = workbook.add_worksheet(ws1.title)
              

                for i in range(1,ws1.max_row + 1):
                    if ws1.cell(row=i, column=6).value != None:
                        ws1.cell(row=i, column=7).value = ws1.cell(row=i, column=6).value

                        textoModificado = re.sub(r"(?<=\w)(<br\/>)(?=\w)|(?<=\w) (<br\/>)(?=\w)|(?<=\w) (<br\/>) (?=\w)|(?<=\w)(<br\/>) (?=\w)|(?<=\w) (<br\/>)(?=\+)|(<br\/>)(?=\()|(?<=\))(<br\/>)",
                                                "<LIONBRIDGE-br/>", str(ws1.cell(row=i, column=7).value))
                        ws1.cell(row=i, column=7).value = textoModificado
                    
                    for j in range(1, mc + 2):                     
                     
                        hoja.write(i-1, j-1, ws1.cell(row=i, column=j).value)
                hoja.set_column('A:F', None, None, {'hidden': 1})
                hoja.set_row(0, None, None, {'hidden': True})
                hoja.write(0, 5, 'Text')
                hoja.write(0, 6, 'Translation')
                hoja.write(0, 7, 'BackTranslation')
                hoja.write(0, 8, 'ComparativeReview')           
                try:
                    workbook.close()
                    self.mensajeretornado = ("Successfully Done!")
                except xlsxwriter.exceptions.FileCreateError as e:
                    self.mensajeretornado = "Error Close Current Prepared File"

        elif self.opcion == 2:

            # Create an new Excel file xlsxwriter and add a worksheet.
            workbook = xlsxwriter.Workbook('PrepFiles\\ '+basename)
           
            for ws1 in wb1.worksheets:
                numero += 1
                tab = str(numero)
                hoja = workbook.add_worksheet(ws1.title)
               
                for i in range(1, ws1.max_row + 1):

                    if ws1.cell(row=i, column=6).fill.start_color.index != '00000000':

                        ws1.cell(row=i, column=7).value = copy(
                            ws1.cell(row=i, column=6).value)
                        textoModificado = re.sub(r"(?<=\w)(<br\/>)(?=\w)|(?<=\w) (<br\/>)(?=\w)|(?<=\w) (<br\/>) (?=\w)|(?<=\w)(<br\/>) (?=\w)|(?<=\w) (<br\/>)(?=\+)|(<br\/>)(?=\()|(?<=\))(<br\/>)",
                                                 "<LIONBRIDGE-br/>", str(ws1.cell(row=i, column=7).value))
                        ws1.cell(row=i, column=7).value = textoModificado

                    else:

                        ws1.cell(row=i, column=7).value = None

                    for j in range(1, mc + 2):

                        if ws1.cell(row=i, column=6).value == None and ws1.cell(row=i, column=7).value != None:
                            ws1.cell(row=i, column=7).value = " "
                        else:
                            hoja.write(
                                i-1, j-1, ws1.cell(row=i, column=j).value)
                hoja.set_column(
                                'A:F', None, None, {'hidden': 1})
                hoja.set_row(0, None, None, {'hidden': True})
                hoja.write(0, 5, 'Text')
                hoja.write(0, 6, 'Translation')
                hoja.write(0, 7, 'BackTranslation')
                hoja.write(0, 8, 'ComparativeReview')         

            try:

                workbook.close()
                self.mensajeretornado = ("Successfully Done!")
            except xlsxwriter.exceptions.FileCreateError as e:
                self.mensajeretornado = "Error Close Current Prepared File"

        elif self.opcion == 3:

            for ws1 in wb1.worksheets:
                numero += 1
                tab = str(numero)
                # Create an new Excel file xlsxwriter and add a worksheet.
                workbook = xlsxwriter.Workbook(
                    'PrepFiles\\ '+ws1.title+".xlsx")
                hoja = workbook.add_worksheet(ws1.title)
               

                for i in range(1, ws1.max_row + 1):

                    if ws1.cell(row=i, column=6).fill.start_color.index != '00000000':

                        ws1.cell(row=i, column=7).value = copy(
                            ws1.cell(row=i, column=6).value)
                        textoModificado = re.sub(r"(?<=\w)(<br\/>)(?=\w)|(?<=\w) (<br\/>)(?=\w)|(?<=\w) (<br\/>) (?=\w)|(?<=\w)(<br\/>) (?=\w)|(?<=\w) (<br\/>)(?=\+)|(<br\/>)(?=\()|(?<=\))(<br\/>)",
                                                 "<LIONBRIDGE-br/>", str(ws1.cell(row=i, column=7).value))
                        ws1.cell(row=i, column=7).value = textoModificado

                    else:

                        ws1.cell(row=i, column=7).value = None

                    for j in range(1, mc + 2):

                        if ws1.cell(row=i, column=6).value == None and ws1.cell(row=i, column=7).value != None:
                            ws1.cell(row=i, column=7).value = " "
                        else:
                            hoja.write(
                                i-1, j-1, ws1.cell(row=i, column=j).value)
                         
                hoja.set_column(
                                'A:F', None, None, {'hidden': 1})
                hoja.set_row(0, None, None, {'hidden': True})
                hoja.write(0, 5, 'Text')
                hoja.write(0, 6, 'Translation')
                hoja.write(0, 7, 'BackTranslation')
                hoja.write(0, 8, 'ComparativeReview')
                try:

                    workbook.close()
                    self.mensajeretornado = ("Successfully Done!")
                except xlsxwriter.exceptions.FileCreateError as e:
                    self.mensajeretornado = ("Successfully Done!")

    def migration(self):
        numero = 0

        newpath = r'PrepFiles'
        if not os.path.exists(newpath):
            os.makedirs(newpath)
        wb1 = xl.load_workbook(self.path)
        ws1 = wb1.active

        rangeselected = []
        basename=os.path.basename(self.path)
        mr = ws1.max_row
        mc = ws1.max_column

        # FTBT DEFAULT PREP
        if self.opcion == 0:

            # Create an new Excel file xlsxwriter and add a worksheet.
            workbook = xlsxwriter.Workbook('PrepFiles\\ '+basename)

            for ws1 in wb1.worksheets:
                numero += 1
                tab = str(numero)
                hoja = workbook.add_worksheet(ws1.title)
                hoja.set_column(
                                'A:F', None, None, {'hidden': 1})

                for i in range(1, ws1.max_row + 1):
                    if ws1.cell(row=i, column=6).value != None:
                        ws1.cell(row=i, column=7).value = ws1.cell(row=i, column=6).value


                        textoModificado = re.sub(r"(?<=\w)(<br\/>)(?=\w)|(?<=\w) (<br\/>)(?=\w)|(?<=\w) (<br\/>) (?=\w)|(?<=\w)(<br\/>) (?=\w)|(<br\/>)(?=\()|(?<=\))(<br\/>)",
                                             "<LIONBRIDGE-br/>", str(ws1.cell(row=i, column=7).value))
                        ws1.cell(row=i, column=7).value = textoModificado

                    for j in range(1, mc + 2):
                      
                            hoja.write(
                                i-1, j-1, ws1.cell(row=i, column=j).value)

                hoja.set_row(0, None, None, {'hidden': True})
                hoja.write(0, 5, 'Text')
                hoja.write(0, 6, 'Translation')
                hoja.write(0, 7, 'BackTranslation')
            try:

                workbook.close()
                self.mensajeretornado = ("Successfully Done!")
            except xlsxwriter.exceptions.FileCreateError as e:
                self.mensajeretornado = "Error Close Current Prepared File"

        elif self.opcion == 1:

            # Create an new Excel file xlsxwriter and add a worksheet.

            for ws1 in wb1.worksheets:
                numero += 1
                tab = str(numero)
                workbook = xlsxwriter.Workbook(
                    'PrepFiles\\ '+ws1.title+".xlsx")
                hoja = workbook.add_worksheet(ws1.title)
            

                for i in range(1, ws1.max_row + 1):
                    if ws1.cell(row=i, column=6).value != None:
                        ws1.cell(row=i, column=7).value = ws1.cell(row=i, column=6).value
                  

                        textoModificado = re.sub(r"(?<=\w)(<br\/>)(?=\w)|(?<=\w) (<br\/>)(?=\w)|(?<=\w) (<br\/>) (?=\w)|(?<=\w)(<br\/>) (?=\w)|(<br\/>)(?=\()|(?<=\))(<br\/>)",
                                                "<LIONBRIDGE-br/>", str(ws1.cell(row=i, column=7).value))
                        ws1.cell(row=i, column=7).value = textoModificado
                    for j in range(1, mc + 2):
                      
                            hoja.write(
                                i-1, j-1, ws1.cell(row=i, column=j).value)
                           
                hoja.set_column(
                                'A:F', None, None, {'hidden': 1})
                hoja.set_row(0, None, None, {'hidden': True})
                hoja.write(0, 5, 'Text')
                hoja.write(0, 6, 'Translation')
                hoja.write(0, 7, 'BackTranslation')
                try:
                    workbook.close()
                    self.mensajeretornado = ("Successfully Done!")
                except xlsxwriter.exceptions.FileCreateError as e:
                    self.mensajeretornado = "Error Close Current Prepared File"

        elif self.opcion == 2:

            # Create an new Excel file xlsxwriter and add a worksheet.
            workbook = xlsxwriter.Workbook('PrepFiles\\ '+basename)

            for ws1 in wb1.worksheets:
                numero += 1
                tab = str(numero)
                hoja = workbook.add_worksheet(ws1.title)

                for i in range(1, ws1.max_row+1 ):

                    if ws1.cell(row=i, column=6).fill.start_color.index != '00000000':

                        ws1.cell(row=i, column=7).value = copy(
                            ws1.cell(row=i, column=6).value)
                        textoModificado = re.sub(r"(?<=\w)(<br\/>)(?=\w)|(?<=\w) (<br\/>)(?=\w)|(?<=\w) (<br\/>) (?=\w)|(?<=\w)(<br\/>) (?=\w)|(<br\/>)(?=\()|(?<=\))(<br\/>)",
                                                 "<LIONBRIDGE-br/>", str(ws1.cell(row=i, column=7).value))
                        ws1.cell(row=i, column=7).value = textoModificado

                    else:

                        ws1.cell(row=i, column=7).value = None

                    for j in range(1, mc + 2):
                        if ws1.cell(row=i, column=6).value == None and ws1.cell(row=i, column=7).value != None:
                            ws1.cell(row=i, column=7).value = " "
                        else:

                            hoja.write(
                                i-1, j-1, ws1.cell(row=i, column=j).value)
                          
                hoja.set_column(
                                'A:F', None, None, {'hidden': 1})
                hoja.set_row(0, None, None, {'hidden': True})
                hoja.write(0, 5, 'Text')
                hoja.write(0, 6, 'Translation')
                hoja.write(0, 7, 'BackTranslation')
            try:

                workbook.close()
                self.mensajeretornado = ("Successfully Done!")
            except xlsxwriter.exceptions.FileCreateError as e:
                self.mensajeretornado = "Error Close Current Prepared File"

        elif self.opcion == 3:
            print('Processing... ')

            for ws1 in wb1.worksheets:
                numero += 1
                tab = str(numero)
                # Create an new Excel file xlsxwriter and add a worksheet.
                workbook = xlsxwriter.Workbook(
                    'PrepFiles\\ '+ws1.title+".xlsx")
                hoja = workbook.add_worksheet(ws1.title)

                for i in range(1,ws1.max_row+1):

                    if ws1.cell(row=i, column=6).fill.start_color.index != '00000000':

                        ws1.cell(row=i, column=7).value = copy(
                            ws1.cell(row=i, column=6).value)
                        textoModificado = re.sub(r"(?<=\w)(<br\/>)(?=\w)|(?<=\w) (<br\/>)(?=\w)|(?<=\w) (<br\/>) (?=\w)|(?<=\w)(<br\/>) (?=\w)|(<br\/>)(?=\()|(?<=\))(<br\/>)",
                                                 "<LIONBRIDGE-br/>", str(ws1.cell(row=i, column=7).value))
                        ws1.cell(row=i, column=7).value = textoModificado

                    else:

                        ws1.cell(row=i, column=7).value = None

                    for j in range(1, mc +2 ):
                        if ws1.cell(row=i, column=6).value == None and ws1.cell(row=i, column=7).value != None:
                            ws1.cell(row=i, column=7).value = " "
                        else:
                            hoja.write(
                                i-1, j-1, ws1.cell(row=i, column=j).value)
                hoja.set_column(
                                'A:F', None, None, {'hidden': 1})
                hoja.set_row(0, None, None, {'hidden': True})
                hoja.write(0, 5, 'Text')
                hoja.write(0, 6, 'Translation')
                hoja.write(0, 7, 'BackTranslation')        

                try:

                    workbook.close()
                    self.mensajeretornado = ("Successfully Done!")
                except xlsxwriter.exceptions.FileCreateError as e:
                    self.mensajeretornado = ("Successfully Done!")

    def TMSftbt(self):
        numero = 0

        newpath = r'PrepFiles'
        if not os.path.exists(newpath):
            os.makedirs(newpath)
        wb1 = xl.load_workbook(self.path)
        ws1 = wb1.active

        rangeselected = []

        mr = ws1.max_row
        mc = ws1.max_column

        if self.opcion == 0:

            # Create an new Excel file xlsxwriter and add a worksheet.

            for ws1 in wb1.worksheets:
                numero += 1
                tab = str(numero)
                workbook = xlsxwriter.Workbook(
                    'PrepFiles\\ '+ws1.title+".xlsx")
                hoja = workbook.add_worksheet(ws1.title)
                
                for i in range(1, ws1.max_row+1):
                
                    if ws1.cell(row=i, column=6).value != None:
                        ws1.cell(row=i, column=7).value = ws1.cell(row=i, column=6).value
                    
                    else:
                       
                        ws1.cell(row=i, column=7).value=" "
                        ws1.cell(row=i, column=6).value=" "
                    textoModificado = re.sub(r"(?<=\w)(<br\/>)(?=\w)|(?<=\w) (<br\/>)(?=\w)|(?<=\w) (<br\/>) (?=\w)|(?<=\w)(<br\/>) (?=\w)|(<br\/>)(?=\()|(?<=\))(<br\/>)",
                                             "<LIONBRIDGE-br/>", str(ws1.cell(row=i, column=7).value))
                    ws1.cell(row=i, column=7).value = textoModificado
                  

                    for j in range(1, mc + 2):
                       
                            hoja.write(
                                i-1, j-1, ws1.cell(row=i, column=j).value)

                hoja.write(0, 5, 'Text')
                hoja.write(0, 6, 'Translation')
                hoja.write(0, 7, 'BackTranslation')
                hoja.write(0, 8, 'ComparativeReview')          

                try:
                    workbook.close()
                    self.mensajeretornado = ("Successfully Done!")
                except xlsxwriter.exceptions.FileCreateError as e:
                    self.mensajeretornado = "Error Close Current Prepared File"

    def TMSmigration(self):
        numero = 0

        newpath = r'PrepFiles'
        if not os.path.exists(newpath):
            os.makedirs(newpath)
        wb1 = xl.load_workbook(self.path)
        ws1 = wb1.active

        rangeselected = []

        mr = ws1.max_row
        mc = ws1.max_column

        if self.opcion == 0:

            # Create an new Excel file xlsxwriter and add a worksheet.

            for ws1 in wb1.worksheets:
                numero += 1
                tab = str(numero)
                workbook = xlsxwriter.Workbook(
                    'PrepFiles\\ '+ws1.title+".xlsx")
                hoja = workbook.add_worksheet(ws1.title)
               
                for i in range(1, ws1.max_row+1):
                    if ws1.cell(row=i, column=6).value != None:
                        ws1.cell(row=i, column=7).value = ws1.cell(row=i, column=6).value
                    else:
                        ws1.cell(row=i, column=7).value=" "
                        ws1.cell(row=i, column=6).value=" "
                    textoModificado = re.sub(r"(?<=\w)(<br\/>)(?=\w)|(?<=\w) (<br\/>)(?=\w)|(?<=\w) (<br\/>) (?=\w)|(?<=\w)(<br\/>) (?=\w)|(<br\/>)(?=\()|(?<=\))(<br\/>)",
                                             "<LIONBRIDGE-br/>", str(ws1.cell(row=i, column=7).value))
                    ws1.cell(row=i, column=7).value = textoModificado

                    for j in range(1,ws1.max_column+2):
                       
                            hoja.write(
                                i-1, j-1, ws1.cell(row=i, column=j).value)

                           
                #hoja.write(0, 5, 'Text')
                hoja.write(0, 6, 'Translation')
            
                try:
                    workbook.close()
                    self.mensajeretornado = ("Successfully Done!")
                except xlsxwriter.exceptions.FileCreateError as e:
                    self.mensajeretornado = "Error Close Current Prepared File"
